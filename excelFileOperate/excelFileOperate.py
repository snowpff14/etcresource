import pandas as pd
import configparser
from pathlib import Path
from datetime import datetime
import numpy as np
import csv
import openpyxl as pyxl
from openpyxl.styles.borders import Border,Side
from logger import LoggerObj 
import os
from pathlib import Path

iniFile=configparser.ConfigParser()
iniFile.read("resources/appConfig.ini","UTF-8")

class InputDataError(Exception):
    def __init__(self,message):
        self.message=message

class DataInsertAndDelete():
  # CSVファイルをソートして取得する
    def getSortedCSVFiles(self,filePath):
        targetFilePath=Path(filePath)
        
        return sorted(targetFilePath.glob('*.csv'))

    # ファイルのマージを行う
    def insertCell(self):
        logObj=LoggerObj('fileMerge')
        log=logObj.createLog('fileMerge')

        log.info('処理開始')
        # basedata=pd.read_excel('基準データ.xlsx')
        baseBook=pyxl.load_workbook(filename ='input/基準データ.xlsx')
        baseSheet=baseBook['データ']

        inputWorkbook=pd.ExcelFile('input/定義ファイル.xlsx')
        addColumnSheet=inputWorkbook.parse('追加列')

        side = Side(style='thin', color='000000')
        blackBorder = Border(top=side,left=side,right=side,bottom=side)
        # 列の追加を先に処理する
        addColumnInfos=np.asarray(addColumnSheet)
        for columninfo in addColumnInfos:
            addColumnName=columninfo[1]
            addColumnPosition=columninfo[2]
            addColumnData=columninfo[3]
            addColmunDatas=inputWorkbook.parse(addColumnData)
            addDatas=np.asarray(addColmunDatas)
            baseSheet.insert_cols(addColumnPosition,1)
            baseSheet.cell(column=addColumnPosition,row=2,value=addColumnName).border=blackBorder
            for i,data in enumerate (addDatas):
                baseSheet.cell(column=addColumnPosition,row=i+3,value=data[1]).border=blackBorder
        date=datetime.now().strftime("%Y%m%d%H%M%S")
        outputDir='output/'+date

        os.makedirs(outputDir,exist_ok=True)

        baseBook.save(outputDir+'/addAfterSheet.xlsx')
        return outputDir+'/addAfterSheet.xlsx'

    # あとはここで引数で変換したファイルを受け取って行を削除
    def deleteRow(self,inputFile):
        logObj=LoggerObj('fileMerge')
        log=logObj.createLog('fileMerge')

        log.info('処理開始')
        # basedata=pd.read_excel('基準データ.xlsx')
        baseBook=pyxl.load_workbook(filename =inputFile)
        baseSheet=baseBook['データ']

        inputWorkbook=pd.ExcelFile('input/定義ファイル.xlsx')
        deleteRowSheet=inputWorkbook.parse('削除行')

        inputWorkbook=pd.ExcelFile(inputFile)
        targetDeleteRowSheet=inputWorkbook.parse('データ',header=1)

        # 列の追加を先に処理する
        addColumnInfos=np.asarray(deleteRowSheet)
        startIndex=3
        for columninfo in addColumnInfos:
            delTargetColumnName=columninfo[1]
            delRowCondition=columninfo[2]
            deleteRows=targetDeleteRowSheet.query(delTargetColumnName+'=="'+delRowCondition+'"')
            deleteIndexs=deleteRows.index

            log.info(deleteIndexs)
            for delcount,delindex in enumerate(deleteIndexs):
                
                baseSheet.delete_rows(startIndex+delindex-delcount)
        
        folder=os.path.dirname(inputFile)

        baseBook.save(folder+'/deleteAfterSheet.xlsx')
        log.info('処理終了')

if __name__ == "__main__":
    dataInsertAndDelete =DataInsertAndDelete()
    resFile=dataInsertAndDelete.insertCell()
    dataInsertAndDelete.deleteRow(resFile)


