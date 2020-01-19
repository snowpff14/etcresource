import configparser
import os
import time
from datetime import datetime

from utils.logger import LoggerObj

import numpy as np
import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import Select, WebDriverWait

from openpyxl.styles.borders import Border, Side
import openpyxl as excel
from openpyxl.utils import get_column_letter


class WebExecuteBase():

    # 出力結果を格納するパス
    # 複数ファイルを設定できるように配列
    resultFilePath = []

    # テスト用かどうかの呼び分けを行う mode=1のときテスト用 testBaseFilePathはテスト用の設定ファイルの位置
    def init(self,
             iniFile,
             mode=0,
             filePaths=[
                 'resources/appConfig.ini'
             ]):
        if mode == 0:
            # 共通で使う項目を定義 設定ファイルを読み込む
            iniFile.read(filePaths[0], "UTF-8")
        elif mode == 1:
            # テスト時はこちらを読み込む
            iniFile.read(filePaths[0], "UTF-8")
        else:
            log = LoggerObj()
            log.error('mode不正')

    # 出力結果のディレクトリを格納単体テストの時にこのパスを受け取って比較を行う
    def setResultPath(self, filePath):
        self.resultFilePath = filePath

    # 出力結果が入っているパスを返す
    def getResultPath(self):
        return self.resultFilePath

    # webdriver用の設定をもろもろ行って返却する
    def createWebDriver(self, iniFile):

        driver=webdriver.Chrome('C:/webdrivers/chromedriver.exe')
        return driver

    # 黒い罫線を返すエクセルのセルに黒で罫線を引きたいときに使用する
    def blackBorderLine(self):
        border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'))

        return border

    # エラー情報をまとめて出力する
    # ディレクトリに付与する年月日時分、CSVのヘッダー、エラー情報をまとめている辞書を引数として渡す
    def outPutInfoCSV(self,iniFile,colums,infoDict,csvFileName='OutPutInfo',targetDateHourMinute=datetime.now().strftime("%Y%m%d%H%M")):
        infoDirectry = iniFile.get(
            "files", "infoDirectry") + targetDateHourMinute
        os.makedirs(infoDirectry, exist_ok=True)

        infoDataFrame=pd.DataFrame(columns=colums)
        indexCount=1
        for key,dataList in infoDict.items():
            for data in dataList:
                # キーの情報、各種情報のリストの形でまとめる
                infoList=[key]
                infoList.extend(data)
                infoSeries=pd.Series(infoList,index=colums,name=indexCount)
                indexCount=indexCount+1
                infoDataFrame=infoDataFrame.append(infoSeries)
        
        errorOutputFileName=infoDirectry+'/{0}.csv'.format(csvFileName)

        infoDataFrame.to_csv(errorOutputFileName,encoding='cp932')
        return errorOutputFileName

    # エラー情報をまとめて出力する
    # ディレクトリに付与する年月日時分、エクセルのヘッダー、エラー情報をまとめている辞書、各セルの幅を格納したリストを引数として渡す
    def outPutInfoExcel(self,iniFile,colums,infoDict,columSize,csvFileName='OutPutInfo',targetDateHourMinute=datetime.now().strftime("%Y%m%d%H%M")):
        infoDirectry = iniFile.get(
            "files", "infoDirectry") + targetDateHourMinute
        os.makedirs(infoDirectry, exist_ok=True)
        infoDataFrame=pd.DataFrame(columns=colums)
        indexCount=1
        for key,dataList in infoDict.items():
            # キーの情報、各種情報のリストの形でまとめる
            for datas in dataList: 
                infoList=[key]
                dList=list(datas)
                infoList.extend(dList)
                infoSeries=pd.Series(infoList,index=colums,name=indexCount)
                indexCount=indexCount+1
                infoDataFrame=infoDataFrame.append(infoSeries)
        
        outputFileName=infoDirectry+'/{0}.xlsx'.format(csvFileName)

        # 結果を一度エクセルに出力
        infoDataFrame.to_excel(outputFileName,encoding='cp932')

        # 罫線を引くようの設定
        border = self.blackBorderLine()

        # ヘッダーの色付け
        fill = excel.styles.PatternFill(patternType='solid',
                                        fgColor='87ceeb', bgColor='87ceeb')


        # フォーマットを整える
        targetBook=excel.load_workbook(outputFileName)
        sheet=targetBook.active
        columLength=len(colums)
        
        for index in range(0,columLength+1):
            # 呼び出し元でカラムのサイズを指定して渡す
            sheet.column_dimensions[get_column_letter(index+1)].width=columSize[index]

        for index,row in enumerate(sheet.rows):
            for cell in row:
                cell.border=border
                if index==0:
                    cell.fill=fill

        targetBook.save(filename=outputFileName)
        return outputFileName


# メイン処理
if __name__ == "__main__":
    print('webBase')

